from openpyxl import load_workbook
import pyautogui

print(pyautogui.position())


path = r'C:\Users\Carvalho-Admin\Downloads\EFD-C 02-22.xlsx'

wb = load_workbook(path)

ws = wb.active

codigos = []
for i in range(22,238):

    codigo = ws.cell(i,1).value

    codigos.append(codigo)

print(codigos)

print(len(codigos))