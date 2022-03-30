import os
import re
import pyautogui as pg
from time import sleep
# from Funcoes.Reconhecimento_de_CNPJ import buscando_posicao_cnpj
from openpyxl import load_workbook

class EFD_Contribuicoes:

    def renomeando_arquivos(self):

        path = r'\\srvcg01\Compartilhados\FISCAL\0 - FISCAL\Membros\Paulo\SuperNova Rotinas\Envio de Escrituracoes EFD - Contribuicoes\2022'
        anos = os.listdir(path)

        for ano in anos:

            path_dois = rf'{path}\{ano}'

            empresas = os.listdir(path_dois)

            for empresa in empresas:

                path_tres = fr'{path_dois}\{empresa}'
                escrituracoes = os.listdir(path_tres)

                for escrituracao in escrituracoes:

                    if 'Contribuições' in escrituracao:
                        try:
                            os.rename(fr'{path_tres}\{escrituracao}',fr'{path_tres}\SPED EFD Contribuicoes')
                            escrituracao = 'SPED EFD Contribuicoes'
                        except FileExistsError:
                            pass

                    arquivos_importacao_das_escrituracoes =  os.listdir(fr'{path_tres}\{escrituracao}')

                    for arquivo in arquivos_importacao_das_escrituracoes:
                        print(arquivo)
                        if 'Contribuições' in arquivo:
                            arquivo_ajustado = arquivo.replace('ç','c').replace('õ','o')
                            print(arquivo)
                            os.rename(fr'{path_tres}\{escrituracao}\{arquivo}',fr'{path_tres}\{escrituracao}\{arquivo_ajustado}')

    # def inserindo_dados_0120(self,path,mes,ano):
    #
    #     if len(str(mes)) == 1:
    #         mes = "0" + str(mes)
    #
    #     competencia = f'{mes}{str(ano)}'
    #
    #     with open(path,'r+') as arquivo:
    #
    #         novoArquivo = []
    #
    #         for registro in arquivo:
    #
    #             novoArquivo.append(registro)
    #
    #         print(novoArquivo)
    #         strings_0110 = []
    #         for arquivo in novoArquivo:
    #             padrao = re.compile("\|0110\|")
    #
    #             be = padrao.search(arquivo)
    #             try:
    #                 print(be.group())
    #                 strings_0110.append(arquivo)
    #             except AttributeError:
    #                 pass
    #         posicoes_0110 = []
    #         for string in strings_0110:
    #             posicoes_0110.append(novoArquivo.index(string))
    #         lista_insercoes = [f'|0120|{competencia}|04|\n','|9900|0120|1|\n']
    #         for v,posicao in enumerate(posicoes_0110):
    #             print()
    #             novoArquivo.insert(posicao + 1,lista_insercoes[v])
    #
    #
    #
    #     with open(path, 'w+') as arquivo:
    #         arquivo.writelines(novoArquivo)
    #
    # def corrigindo_linhas(self,path):
    #     #Path é o caminho até o xml
    #
    #
    #     padrao_1 = re.compile('\|0990\|\d\|')
    #     padrao_2 = re.compile('\|99\d\d\|\d\d\|')
    #
    #     #Abrindo arquivo
    #     with open(path, 'r+') as arquivo:
    #         novoArquivo = []
    #         for registro in arquivo:
    #             novoArquivo.append(registro)
    #
    #
    #         padrao1 = []
    #         padrao2 = []
    #         for arquivo in novoArquivo:
    #
    #             a = padrao_1.search(arquivo)
    #
    #             if a != None:
    #                 padrao1.append(arquivo)
    #                 print(arquivo)
    #                 break
    #
    #         for arquivo in novoArquivo:
    #             b = padrao_2.search(arquivo)
    #
    #             if b != None:
    #                 padrao2.append(arquivo)
    #
    #         posicao_9001_0 = novoArquivo.index("|9001|0|\n")
    #         print(posicao_9001_0)
    #
    #
    #         posicoes = []
    #
    #
    #         for arquivo in padrao2:
    #
    #             posicoes.append(novoArquivo.index(arquivo))
    #
    #         for arquivo in padrao1:
    #             posicoes.append(novoArquivo.index(arquivo))
    #
    #         print(posicoes)
    #         novoArquivo[posicoes[0]] = f'|9900|9900|{posicoes[1] - posicao_9001_0 - 1}|\n'
    #         novoArquivo[posicoes[1]] = f'|9990|{posicoes[2] + 1 - posicao_9001_0}|\n'
    #         novoArquivo[posicoes[2]] = f'|9999|{posicoes[2] + 1}|\n'
    #         novoArquivo[posicoes[3]] = f'|0990|{str(posicoes[3] + 1)}|\n'
    #
    #
    #         with open(path, 'w+') as arquivo:
    #             arquivo.writelines(novoArquivo)

    def importando(self,empresa,mes,ano):
        if len(str(mes)) <= 1:
            mes = "0" + str(mes)
        print(mes)
        competencia = f'{str(ano)}-{mes}'

        qtd_zeros = 6 - len(str(empresa))

        empresa = '0' * qtd_zeros + str(empresa)
        arquivo =  fr"\\srvcg01\Compartilhados\FISCAL\0 - FISCAL\Membros\Paulo\SuperNova Rotinas\Envio de Escrituracoes EFD - Contribuicoes\2022\{mes}\{empresa}\SPED EFD Contribuicoes\SPED EFD Contribuicoes '{empresa}' ({competencia}).txt"

        print(arquivo)


        if os.path.exists(arquivo) == False:
            print('O arquivo não existe, gere-o ou renomeie-o')

            return None

        'Supondo que o SPED EFD Contribuições esteja aberto'
        pg.sleep(2)
        pg.hotkey('ctrl','i')

        pg.sleep(2)
        escrituracao_importada = pg.locateOnScreen(r'Imagens\importarEscrituracao.png',confidence=0.95)

        if escrituracao_importada !=None:

            pg.typewrite(arquivo)
            pg.sleep(0.5)
            pg.press('enter')

            while True:
                jaExiste = pg.locateOnScreen(r'Imagens\jaExiste.png',confidence=0.95)
                importadaSucesso = pg.locateOnScreen(r'Imagens\importadaExito.png',confidence=0.95)



                if jaExiste != None:
                    pg.press('enter')

                    sleep(3)

                    pg.press('enter')


                    'Verificar se o arquivo foi validado com sucesso'
                if importadaSucesso != None:
                    pg.press('enter')

                    resultado = ""
                    while True:
                        validadoSucesso = pg.locateOnScreen(r'Imagens\validadoSucesso.png',confidence=0.95)
                        erroEstrutura = pg.locateOnScreen(r'Imagens\errosImportacao.png',confidence=0.95)
                        contemAvisos = pg.locateOnScreen(r'Imagens\contemAvisos.png',confidence=0.95)
                        contemErros = pg.locateOnScreen(r'Imagens\contemErros.png', confidence=0.95)
                        print('procurando')

                        if validadoSucesso != None:
                            pg.press('enter')
                            sleep(5)

                            print('segundo enter')
                            pg.press('enter')
                            resultado = "Validado com sucesso"
                            break

                        if contemAvisos !=None:
                            print('contem aviso')
                            pg.press('enter')
                            sleep(1)
                            pg.press('enter')
                            resultado = "Validado com sucesso"
                            break

                        if erroEstrutura != None:
                            pg.press('enter')
                            resultado = "Erro na importacao"

                            break

                        if contemErros != None:
                            pg.press('enter')
                            sleep(1)
                            resultado = 'Erro na importacao'
                            break

                    sleep(2)
                    pg.hotkey('ctrl','f')
                    return resultado






        jaExiste = pg.locateOnScreen(r'Imagens\jaExiste.png',confidence=0.95)


    def assinando(self,cnpj,erros_de_assinatura):

        pg.hotkey('ctrl','s')
        sleep(3)

        # posicao = buscando_posicao_cnpj(cnpj)
        #
        # if posicao == None:
        #
        #     pg.press('esc')
        #     return False
        # # se nao tiver na imagem, vai retornar None, se retornar None, clicar pra baixo
        # pg.leftClick(745,494 + ((posicao-1)*15))
        # sleep(1)
        pg.press('down', presses=erros_de_assinatura + 1, interval=0.3)
        sleep(1)
        pg.moveTo(x=20, y=1002)
        pg.press('enter')


        #Processo pra chegar no certitifacdo

        sleep(2)
        pg.press('down',presses=16)
        sleep(1)
        pg.press('tab',presses=4,interval=0.3)
        sleep(0.5)
        pg.press('enter')

        while True:

            validadoSucesso = pg.locateOnScreen(r'Imagens\assinadoSucesso.png',confidence=0.95)
            erro = None

            if validadoSucesso != None:
                sleep(1)
                print('validado com sucesso')
                pg.press('enter')
                return 'Validado'

            if erro != None:


                pass

    def transmitindo(self,cnpj,codigo,nome,erros_de_transmissao):

        pasta_de_salvar = fr'\\srvcg01\Compartilhados\FISCAL\0 - FISCAL\Membros\Paulo\SuperNova Rotinas\Envio de Escrituracoes EFD - Contribuicoes\2022\Recibos - 02\{codigo} - {nome} - EFD Contribuicoes'

        pg.hotkey('ctrl', 't')
        sleep(3)

        # posicao = buscando_posicao_cnpj(cnpj)
        # sleep(1)
        # if posicao == None:
        #     pg.press('esc')
        #     return False
        # # se nao tiver na imagem, vai retornar None, se retornar None, clicar pra baixo
        # pg.leftClick(745, 494 + ((posicao - 1) * 15))

        pg.press('down',presses=erros_de_transmissao + 1,interval=0.3)


        sleep(1)
        pg.moveTo(x=20, y=1002)
        pg.press('enter')


        while True:

            transmitidoSucesso = pg.locateOnScreen(r'Imagens\transmitidoSucesso.png',confidence=0.95)
            arquivoNaoTransmitido = pg.locateOnScreen(r'Imagens\arquivoNaoTransmitido.png',confidence=0.95)


            if transmitidoSucesso != None:

                pg.press('enter')
                sleep(1)


                while True:
                    reciboTransmissao = pg.locateOnScreen(r'Imagens\reciboTransmissao.png', confidence=0.95)
                    print(reciboTransmissao)
                    if reciboTransmissao != None:

                        pg.press('tab')
                        sleep(0.5)
                        pg.press('enter',presses=2,interval=0.5)

                        sleep(2)

                        pg.typewrite(pasta_de_salvar)

                        pg.press('enter')

                        sleep(2)

                        pg.press('esc')

                        break
                return True
            if arquivoNaoTransmitido != None:

                pg.press('enter')
                return False

a = EFD_Contribuicoes()

#a.inserindo_dados_0120(path,2,2022)

#a.corrigindo_linhas(path)


path = r'C:\Users\Carvalho-Admin\Downloads\EFD-C 02-22.xlsx'

wb = load_workbook(path)

ws = wb.active

codigos = []
erros_de_assinatura = ws.cell(3,13).value
erros_na_transmissao = ws.cell(3, 14).value
print(erros_na_transmissao,erros_de_assinatura)

for i in range(88,100):

    codigo = ws.cell(i,1).value
    cnpj = ws.cell(i,4).value
    nome = ws.cell(i,2).value
    print('CNPJ é ' + cnpj)
    importando = a.importando(codigo,2,2022)

    if importando == 'Validado com sucesso':
        assinatura = a.assinando(str(cnpj),erros_de_assinatura)
        if assinatura != False:
            transmissao = a.transmitindo(cnpj,str(codigo),nome,erros_na_transmissao)

            if transmissao == False:
                erros_na_transmissao += 1
            sleep(5)
        else:
            erros_de_assinatura +=1
